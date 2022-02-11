#model
import openpyxl
import os
import datetime
import csv
import PySimpleGUI as sg

def CheckAndAddPath(v):
    #this function checks if filepath is correct and adds path if there is only filename.
    if (":" in v) ==True:
        return v
    else:
        return str(os.path.dirname(__file__))+'/'+v

def ReplaceSpaceSeparatorAndCommaMbank(s):
    #this functions checks fields is mbank sheet, removes unecessary spaces and replacing commas to dots to work properly in xlsx
    if " " in s and s !="  ":
        s1 = s.replace(" ","")
        if s1[0].isnumeric() == True or s1[0] =="-":
            s = s.replace(" ","")
    if "," in s:
        s = s.replace(",",".")    
    try:
        s=float(s)
    except ValueError:
        pass
    return s

#a = ReplaceSpaceSeparatorAndCommaMbank("22,0")
#print(a)

def IfNumberMakeFloat(s):
    #this function changes numbers with + or - to normal floats.
    if " " in s or s =="":
        return s
    elif s[0]=="-" or s[0]=="+":
        s = float(s)
        return s
    else:
        return s



def AddToDictionary(short,category):
    #this function adds short name and cateogry to dictionary
    autofolder = str(os.path.dirname(__file__))
    file_to_open = autofolder + "/dict.csv"
    f = open(file_to_open,"a+")
    f.write (short+";"+category+"\n")       
    f.close()
    return "ok"

def JoinFiles(File1,File2,File3,NameFile):
    #this function joins files into one xlsx file.
    if File1==""and File2=="" and File3==""and NameFile=="":
        File1 = input("first file (csv): ")
        File2 = input("second file csv: ")
        File3 = input("third file xlsx: ")
        NameFile = input("New filename (with .xlsx): ")
    File1 = CheckAndAddPath(File1)
    PathNameFile = str(os.path.dirname(__file__))+'/'+NameFile
    #print (os.path.dirname(__file__))
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title ="file1" 
    f = open(File1,'rt')
    reader = csv.reader(f,delimiter=";")
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            s = cell
            s = ReplaceSpaceSeparatorAndCommaMbank(s)
            ws1.cell(row=(row_index + 1),column = column_index+1).value = s      
    if File2 =="":
        pass
        #print("No second file")
    else:
        File2 = CheckAndAddPath(File2)
        ws2 = wb.create_sheet("file2")
        f = open(File2,'rt')
        reader = csv.reader(f,delimiter=";")
        for row_index, row in enumerate(reader):
            for column_index, cell in enumerate(row):
                s = cell
                s = ReplaceSpaceSeparatorAndCommaMbank(s)
                ws2.cell(row=(row_index + 1),column = column_index+1).value = s 
    #print(File3)
    if File3 =="":
        pass
        #print("No third file")
    else:
        File3 = CheckAndAddPath(File3)
        ws3 = wb.create_sheet("file3")
        f = open(File3,'rt')
        reader = csv.reader(f,delimiter=",")
        for row_index, row in enumerate(reader):
            for column_index, cell in enumerate(row):
                s = cell
                #print (cell)
                s = IfNumberMakeFloat(s)
                ws3.cell(row=(row_index + 1),column = column_index+1).value = s
    wb.save(PathNameFile)
    return "ok"

#JoinFiles()

def SaveCategory(i,desc, cat,dictyn,file,sheet,bank):
    #this functions adds category to xlsx file
    autofolder = str(os.path.dirname(__file__))
    wb = openpyxl.load_workbook(filename = autofolder + '/'+file)
    ws = SelectSheet(wb,sheet)
    if bank == "mb":
        CategoryPrintCell = 'j'+str(i) 
    else:
        CategoryPrintCell = 'm'+str(i)
    ws[CategoryPrintCell].value = cat
    wb.save(filename = autofolder + '/'+file)
    #print(i,cat,dictyn)
    if dictyn == True:
        AddToDictionary((desc[:10]),cat)
        return "okDict"
    else:
        return "ok"
    #save category

def FindFirstRowMbank(ws):
    #this functions finds first row of transaction list in mbank file
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "#Data operacji":
                return cell.row
                #print (cell.row)

def FindLastRowMbank(ws):
    #this functions finds last row of transaction list in mbank file
    for row in ws.iter_rows():
        for cell in row:
            ShortCellValue = str(cell.value)
            if ShortCellValue[:10] == "Niniejszy ":
                return cell.row
                #print (cell.row)       

def SelectSheet(wb,SelectedSheet):
    #this functions gives user in command line chance to sleect sheet he is interested in
    if SelectedSheet == "":
        print('Sheets: ')
        print(wb.sheetnames)
        SelectedSheet = input("Enter sheet name: ")
    #print(wb.sheetnames)
    return wb[SelectedSheet]

def CountCategoriesMB(file,sheet):
    #this function returns first and last row of mbank file
    autofolder = str(os.path.dirname(__file__))
    wb = openpyxl.load_workbook(filename = autofolder + '/'+file)
    ws = SelectSheet(wb,sheet)
    #znaldz pierwszy wiersz tabelki
    HeaderRow = FindFirstRowMbank(ws)
    LastRow = FindLastRowMbank(ws)
    return HeaderRow,LastRow

def CountCategoriesBP(file,sheet):
    #this function returns first and last row of pko file
    autofolder = str(os.path.dirname(__file__))
    wb = openpyxl.load_workbook(filename = autofolder + '/'+file)
    ws = SelectSheet(wb,sheet)
    HeaderRow = 1
    LastRow = ws.max_row+1
    return HeaderRow,LastRow


def SearchForValue(value):
  # this functions searches for cateogry for transaction description in dictionary and returns if found.
  autofolder = str(os.path.dirname(__file__))
  file_to_open = autofolder + "/dict.csv"
  with open(file_to_open) as f:
    d = dict(x.rstrip().split(";", 1) for x in f)    
  if value in d:
    return(d[value])
  else:
    return ("Not found")

def AssignCategoriesMB(i, file,sheet):
    #this function assign categories to transaction in mbank file
    autofolder = str(os.path.dirname(__file__))
    wb = openpyxl.load_workbook(filename = autofolder + '/'+file)
    ws = SelectSheet(wb,sheet)
    NumberOfCell = 'd'+str(i)
    CategoryPrintCell = 'j'+str(i)
    ShortenedDescription = (ws[NumberOfCell].value[:10])
    if not ws[CategoryPrintCell].value:
        #print(ws[CategoryPrintCell].value)
        x = SearchForValue(ShortenedDescription)
        if x == "Not found" :
            #print("nieznaleziono w słowniku")
            #print(ws[NumberOfCell].value)
            TransactionDate = ws['b'+str(i)].value
            if type(ws['b'+str(i)]) == datetime:
                TransactionDate = ws['b'+str(i)].value.strftime("%d-%m-%Y")
            TransactionAmount = str(ws['g'+str(i)].value)
            TransactionDesc = ws['c'+str(i)].value, str(ws['d'+str(i)].value),ws['e'+str(i)].value
            #print(ws["b"]+str(i)).coordinate
            #print("Date:    "+TransactionDate)
            #print("Amount: "+TransactionAmount)
            return TransactionDate, TransactionAmount, TransactionDesc[0],TransactionDesc[1],TransactionDesc[2]
        else:
            #ws[ShortNameCell].value = ShortenedDescription, ws['c'+str(i)].value
            ws[CategoryPrintCell].value = x
            wb.save(filename = autofolder + '/'+file)
            return "assigned"
    else:
        return "assigned"        

def AssignCategoriesBP(i, file,sheet):
    #this function assign categories to transaction in pkobp file
    #open file return wb and ws
    autofolder = str(os.path.dirname(__file__))
    wb = openpyxl.load_workbook(filename = autofolder + '/'+file)
    ws = SelectSheet(wb,sheet)
    #read content
    row = i
    print(i)
    TransactionTypeCell = "c"+str(row)
    #print(ws[TransactionTypeCell].value)
    AmountCell = 'd'+str(row)
    CategoryPrintCell = "m"+str(row)
    if ws[TransactionTypeCell].value == "Płatność kartą":
        hCell = ws['h'+str(row)].value
        Description = hCell[hCell.find('Adres: ')+7:]
    elif ws[TransactionTypeCell].value == "Przelew z rachunku":
        hCell = ws['h'+str(row)].value
        iCell = ws['i'+str(row)].value
        jCell = ws['j'+str(row)].value
        if (iCell.find("Adres odbiorcy")) == -1:
            Description = iCell[7:]+' '+hCell[16:]
        else:
            Description = jCell[7:]
    elif ws[TransactionTypeCell].value == "Przelew na rachunek":
        Description = "wplywy"
    elif ws[TransactionTypeCell].value == "Płatność web - kod mobilny":
        iCell = ws['i'+str(row)].value
        Description = iCell[iCell.find("Adres: ")+7:]
    elif ws[TransactionTypeCell].value == "Zlecenie stałe":
        iCell = ws['i'+str(row)].value
        Description = iCell[7:]
    elif ws[TransactionTypeCell].value == "Spłata kredytu":
        Description = 'kredyt'
    else:
        Description = "inne"
    print(Description)
    ShortenedDescription = Description[:10]
    x = SearchForValue(ShortenedDescription)
    if x == "Not found" :
        TransactionDate = ws['a'+str(row)].value
        if type(ws['a'+str(row)]) == datetime:
            TransactionDate = ws['b'+str(row)].value.strftime("%d-%m-%Y")
        TransactionAmount = str(ws['d'+str(row)].value)
        TransactionDesc = ws[TransactionTypeCell].value,Description,' '
        print(TransactionDate, TransactionAmount, TransactionDesc[0],TransactionDesc[1],TransactionDesc[2])
        return TransactionDate, TransactionAmount, TransactionDesc[0],TransactionDesc[1],TransactionDesc[2]
    else:
        ws[CategoryPrintCell].value = x
        print(x)
        wb.save(filename = autofolder + '/'+file)
        return "assigned"

def testAssignCategoriesBP():
    #this function tests Assignig cateogry BP
    for i in range(2,9):
        AssignCategoriesBP(i,'entertestfile','entertestsheet')


def ReadCategoryFile():
    #this function returns list with category names
    autofolder = str(os.path.dirname(__file__))
    file_to_open = autofolder + "/categories.csv"
    with open(file_to_open) as file:
        catdict = file.readlines()
        catdict = [line.rstrip() for line in catdict]
    return catdict

def SheetSummaryMbank(file,SelectedSheet):
    #this function creates summary for sheet from mbank file +is ready for extenralized categories
    autofolder = str(os.path.dirname(__file__))
    wb = openpyxl.load_workbook(filename = autofolder + '/'+file)
    ws = SelectSheet(wb,SelectedSheet)
    CatDict = ReadCategoryFile()
    HeaderRow = FindFirstRowMbank(ws)
    LastRow = FindLastRowMbank(ws)
    if LastRow == None: #only for test
        LastRow=6
        HeaderRow =1
    LenCatDict = len(CatDict)-1
    if ws["d"+str(LastRow+2)] == "SUMMARY":
        pass
    else:
        ws["d"+str(LastRow+2)] = "SUMMARY"
        for row in range(LastRow+3,LastRow+LenCatDict+4):
            catnr = row -LastRow-3       
            ws["d"+str(row)] = CatDict[catnr]
        ws["d"+str(LastRow+LenCatDict+5)] = "SUM"
    for row in range(LastRow+3,LastRow+LenCatDict+4):
        #print(row)
        ws["e"+str(row)] = "=SUMIF($J$"+str(HeaderRow)+":$J$"+str(LastRow-4)+",D"+str(row)+",$g$"+str(HeaderRow)+":$g$"+str(LastRow-4)+")"
    ws['e'+str(LastRow+LenCatDict+5)] = "=SUM(e"+str(LastRow+3)+":e"+str(LastRow+LenCatDict+2)+")"   
    wb.save(filename = autofolder + '/'+file)
    print("Summary was added to sheet: "+ str(SelectedSheet))

#SheetSummaryMbank('test.xlsx','test')
#SheetSummaryMbank1('test1.xlsx','test')


def SheetSummaryPkoBP(file,SelectedSheet):
    #this function creates summary for sheet from pkobp file
    autofolder = str(os.path.dirname(__file__))
    wb = openpyxl.load_workbook(filename = autofolder + '/'+file)
    ws = SelectSheet(wb,SelectedSheet)
    HeaderRow = 1
    LastRow = ws.max_row
    SummaryRow = LastRow+2
    #print (SummaryRow)
    #print(ws["d"+str(SummaryRow)].value)
    if SummaryRow == None: #only for test
        LastRow=6
        HeaderRow =1
    CatDict = ReadCategoryFile()
    LenCatDict = len(CatDict)-1
    if ws["d"+str(SummaryRow)].value == "SUMMARY":
        pass
        #print ("Summary already in the file")
    else:
        ws["d"+str(LastRow+2)] = "SUMMARY"
        for row in range(LastRow+3,LastRow+LenCatDict+4):
            catnr = row -LastRow-3       
            ws["d"+str(row)] = CatDict[catnr]
        ws["d"+str(LastRow+28)] = "SUM"
        SummaryRow = LastRow+2
        for row in range(LastRow+3,LastRow+LenCatDict+4):
            #print(row)
            ws["e"+str(row)].value = "=SUMIF($m$"+str(HeaderRow)+":$m$"+str(LastRow)+",D"+str(row)+",$d$"+str(HeaderRow)+":$d$"+str(LastRow)+")"
        #print(SummaryRow+26)
        ws['e'+str(LastRow+LenCatDict+5)].value = "=SUM(e"+str(LastRow+3)+":e"+str(LastRow+LenCatDict+2)+")"  
        #print("Summary was added to sheet: "+ str(SelectedSheet)) 
        wb.save(filename = autofolder + '/'+file)    

#SheetSummaryPkoBP('test.xlsx',"test")

def FileSummary(file,SelectedSheet,SummaryNo):
    #this function creates summary for file 
    autofolder = str(os.path.dirname(__file__))
    wb = openpyxl.load_workbook(filename = autofolder + '/'+file)
    CatDict = ReadCategoryFile()
    LenCatDict = len(CatDict)-1
    if 'Summary' in wb.sheetnames:
        ws4 = wb["Summary"]
    else:
        ws4 = wb.create_sheet("Summary")
        LastRow = -1
        CatDict = ReadCategoryFile()
        LenCatDict = len(CatDict)-1
        ws4["a"+str(LastRow+2)] = "SUMMARY"
        for row in range(LastRow+3,LastRow+LenCatDict+4):
            catnr = row -LastRow-3       
            ws4["a"+str(row)] = CatDict[catnr]
        ws4["a"+str(LastRow+LenCatDict+4)] = "SUM" 
    ws = SelectSheet(wb,SelectedSheet)
    for row in ws.iter_rows():
        for cell in row:
            #print(cell)
            ShortCellValue = str(cell.value)
            if ShortCellValue == "SUMMARY":
                SummaryRow = cell.row
                SummaryCol = cell.column
            else:
                pass
                #print('nic')
    #SummaryCell = openpyxl.utils.get_column_letter(SummaryCol+1)+str(SummaryRow+2)
    SummaryNo = int(SummaryNo)+1
    SourceSummaryLetter = openpyxl.utils.get_column_letter(SummaryCol+1)
    TargetSummaryLetter = openpyxl.utils.get_column_letter(SummaryNo)
    #print(SummaryCell)
    #print(SummaryCol)
    #print(SummaryRow)
    ws4[TargetSummaryLetter+"1"] = ws.title
    for x in range(2,LenCatDict+4):
        SummaryCell = SourceSummaryLetter+str(SummaryRow+x-1)
        #print(TargetSummaryLetter+str(x))
        ws4[TargetSummaryLetter+str(x)] = "='"+ws.title+"'!"+SummaryCell
    #print (ws4['b1'].value)
    #print (ws4['b2'].value)
    #print (ws4['b3'].value)
    wb.save(filename = autofolder + '/'+file)
    #print("Summary of selected sheets prepared for file :"+file)    

def SummarySummary(file):
    #this functions creates sum for all expenses in file summary
    autofolder = str(os.path.dirname(__file__))
    wb = openpyxl.load_workbook(filename = autofolder + '/'+file)
    ws = wb["Summary"]
    CatDict = ReadCategoryFile()
    LenCatDict = len(CatDict)-1
    for row in range (2,LenCatDict+2):
        ws['e'+str(row)].value = "=-b"+str(row)+"-c"+str(row)+"-d"+str(row)
    x = str(LenCatDict+2)    
    ws['e'+x].value = "=b"+x+"+C"+x+"+D"+x
    x = str(LenCatDict+3)
    ws['e'+x].value = "=-b"+x+"-c"+x+"-d"+x
    wb.save(filename = autofolder + '/'+file)
    #print ("Summary of all expenses prepared")

#FileSummary('testgru.xlsx','file1',1)
#FileSummary('testgru.xlsx','file2',2)
#FileSummary('testgru.xlsx','file3',3)
#SummarySummary('testgru.xlsx')

#print(JoinAndProcess("","","",""))